import pandas as pd
import numpy as np
import os
import re
from openpyxl import load_workbook
import openpyxl 
import string 
import datetime
import glob
from datetime import date, timedelta, datetime
import datetime


pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)


""" Copy range of cells as a nested list
    Takes: start cell, end cell, and sheet you want to copy from.
"""

class copysheet: 
    def __init__(self, sheetGiving, sheetRecieving, filesave_name):
        self.sheetGiving = sheetGiving
        self.sheetRecieving = sheetRecieving
        self.filesave_name = filesave_name
        self.file_to_paste = openpyxl.load_workbook(self.sheetRecieving) # Add file name
   
    def copyRange(self, startCol, endCol, startRow, endRow, copySheet):
        rangeSelected = []
        # Loops through selected Rows
        for i in range(startRow,endRow + 1,1):
            # Appends the row to a RowSelected list
            rowSelected = []
            for j in range(startCol,endCol+1,1):
                rowSelected.append(copySheet.cell(row = i, column = j).value)
            # Adds the RowSelected List and nests inside the rangeSelected
            rangeSelected.append(rowSelected)
        return rangeSelected

    def pandasCopyRange(self, dataframe):
        startRow = 0
        print(dataframe)
        endRow = len(dataframe)
        startCol = 0
        endCol = len(dataframe.columns)
        rangeSelected = []
        for i in range(startRow,endRow,1):
            rowSelected = []
            for j in range(startCol,endCol,1):
                rowSelected.append(dataframe.iloc[i, j])
            rangeSelected.append(rowSelected)
        return rangeSelected
        
 
    # Paste data from copyRange into template sheet
    def pasteRange(self, startCol, endCol, startRow, endRow, pasteSheet, copiedData):
        countRow = 0
        for i in range(startRow,endRow+1,1):
            countCol = 0
#             for j in range(startCol,endCol+1,1):
#                 if copiedData[countRow][countCol] in ("(null)", 'null', ' ', '', '#VALUE!'):
#                     pasteSheet.cell(row = i, column = j).value = 0
#                 else:
#                     pasteSheet.cell(row = i, column = j).value = copiedData[countRow][countCol]
                    
#                 if isinstance(copiedData[countRow][countCol], datetime.date):
#                     pasteSheet.cell(row = i, column = j).value = copiedData[countRow][countCol].strftime('%m/%d/%Y')

                    
                countCol += 1
            countRow += 1
        

    def col2num(self, col):
        num = 0
        for c in col:
            if c in string.ascii_letters:
                num = num * 26 + 1 + ord(c) - ord('A')
        return num
    
    def date_slash(self, dataframe): 
        for row in range(0, len(dataframe)):
            dataframe.iloc[row,0] = dataframe.iloc[row,0].strftime('%m/%d/%Y')

#     def selectQuery(self, query): 
#         get_dobj = ObjectDBInstance()
#         query_sql = f""" 
#                   {query}    
#                     """
#         rows = get_dobj.do_sql(query_sql) 
#         header = get_dobj.get_result_column_names()
#         items = pd.DataFrame.from_records(rows, columns=header)
            
#         return items
    
    
    def copyPaste(self, sheet_name_to_copy, Start_Column_copy, Start_Row_copy, End_Column_copy, End_Row_copy, sheet_name_to_paste, Start_Column_paste, Start_Row_paste,End_Column_paste, End_Row_paste):
        # File to copy from
        file_to_copy = openpyxl.load_workbook(self.sheetGiving, data_only=True) # Add file name
        copy_sheet = file_to_copy[sheet_name_to_copy] #Add Sheet name
        # cells to be copied
        to_copy = self.copyRange(self.col2num(Start_Column_copy), self.col2num(End_Column_copy), Start_Row_copy, End_Row_copy, copy_sheet)
        # cells to recieve copied cells
        paste_sheet = self.file_to_paste[sheet_name_to_paste] #Add Sheet name
        
        print("Processing...")
        # selectedRange = pandasCopyRange(sheet) 
        pastingRange = self.pasteRange(self.col2num(Start_Column_paste),self.col2num(End_Column_paste),Start_Row_paste,End_Row_paste,paste_sheet,to_copy)

        print("Range copied and pasted!")
        
    def pandasCopyPaste(self,Start_Column_paste, Start_Row_paste, sheet_name_to_paste, dataframe):
        # cells to be copied
        df = self.pandasCopyRange(dataframe)
        # cells to recieve copied cells
        paste_sheet = self.file_to_paste[sheet_name_to_paste] #Add Sheet name
        End_Column_paste = int(self.col2num(Start_Column_paste)) + int(len(dataframe.columns)) -1
        End_Row_paste = Start_Row_paste + len(df) -1
        print("Processing...")
        # selectedRange = pandasCopyRange(sheet) 
        # print(self.col2num(Start_Column_paste), End_Column_paste, Start_Row_paste, End_Row_paste)
        pastingRange = self.pasteRange(self.col2num(Start_Column_paste),End_Column_paste,Start_Row_paste,End_Row_paste,paste_sheet,df) 

        print("Range copied and pasted to %s!" % (sheet_name_to_paste)) 
    
    
    def writeToCell(self, sheet_name_to_paste, columnLetter , rowNum, val): 
        pasteSheet = self.file_to_paste[sheet_name_to_paste] 
        pasteSheet.cell(row = rowNum, column = self.col2num(columnLetter)).value = val
        print("cell write to %s successful" % (sheet_name_to_paste)) 
            
      
    def save(self):
        self.file_to_paste.save(self.filesave_name)
        print("file saved!")

